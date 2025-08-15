from openpyxl import load_workbook
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler

# --------------------- 数据读取与预处理 ---------------------
# 读取Excel数据
workbook = load_workbook("D:/gold.xlsx")
sheet = workbook["黄金期货历史数据 (1)"]
Xl_range = sheet["E2:E400"]  # 左边界列
Xr_range = sheet["D2:D400"]  # 右边界列

# 提取原始数据
X_L = [cell[0].value for cell in Xl_range]
X_R = [cell[0].value for cell in Xr_range]


# 计算差分序列 ΔX
def compute_delta(X):
    return [X[i] - X[i - 1] for i in range(1, len(X))]


Delta_X_L = compute_delta(X_L)
Delta_X_R = compute_delta(X_R)

# 数据标准化
scaler = StandardScaler()
Y = np.column_stack((Delta_X_L, Delta_X_R))
Y_scaled = scaler.fit_transform(Y)  # 标准化后的差分序列

# 构造二维矩阵列表
Y_list = [np.array([[y[0]], [y[1]]]) for y in Y_scaled]
qt = (Y_scaled[:, 0] + Y_scaled[:, 1]) / 2  # 标准化后的分位点

# --------------------- 模型参数与结构定义 ---------------------
K = np.array([[1, -1], [-1, 5]])  # 权重矩阵
gamma_list = np.quantile(qt, np.linspace(0, 1, 100), method='nearest')

# 构建滞后项矩阵Z（修复索引错误）
Z = []
for i in range(1, len(Y_scaled)):  # 正确滞后索引: i-1 ∈ [0, T-2]
    z_i = np.array([
        [1, -0.5, Y_scaled[i - 1, 0]],  # 滞后一期左差分
        [1, 0.5, Y_scaled[i - 1, 1]]  # 滞后一期右差分
    ])
    Z.append(z_i)
# --------------------- 参数优化主循环 ---------------------
best_DK = float('inf')
best_gamma = None
best_beta1 = None
best_beta2 = None

# 遍历分位点（可并行优化）
for j in range(5, len(gamma_list) - 4):
    gamma = gamma_list[j]

    # 划分区间
    beta1_up, beta1_down = np.zeros((3, 1)), np.zeros((3, 3))
    beta2_up, beta2_down = np.zeros((3, 1)), np.zeros((3, 3))

    for i in range(len(Z)):  # Z长度=T-2
        Z_i = Z[i]
        Y_next = Y_list[i + 1]  # 对应Y[i+1]

        # 计算DK项
        DK_up = Z_i.T @ K @ Y_next
        DK_down = Z_i.T @ K @ Z_i

        if qt[i] < gamma:
            beta1_up += DK_up
            beta1_down += DK_down
        else:
            beta2_up += DK_up
            beta2_down += DK_down

    # 正则化矩阵求逆（防止奇异）
    reg = 1e-5 * np.eye(3)
    try:
        beta1 = np.linalg.inv(beta1_down + reg) @ beta1_up
        beta2 = np.linalg.inv(beta2_down + reg) @ beta2_up
    except np.linalg.LinAlgError:
        continue  # 跳过不可逆情况

    # 计算总残差DK
    DK = 0
    for i in range(len(Z)):
        Z_i = Z[i]
        Y_real = Y_list[i + 1]
        if qt[i] < gamma:
            Y_hat = Z_i @ beta1
        else:
            Y_hat = Z_i @ beta2
        DK += (Y_hat - Y_real).T @ K @ (Y_hat - Y_real)

    # 更新最优参数
    if DK < best_DK:
        best_DK = DK.item()
        best_gamma = gamma
        best_beta1 = beta1
        best_beta2 = beta2

# --------------------- 预测与结果还原 ---------------------
# 生成预测差分
Y_hat_scaled = []
for i in range(len(Z)):
    Z_i = Z[i]
    if qt[i] < best_gamma:
        Y_hat = Z_i @ best_beta1
    else:
        Y_hat = Z_i @ best_beta2
    Y_hat_scaled.append(Y_hat)

# 逆标准化差分预测
Y_hat = scaler.inverse_transform(
    np.concatenate([y.reshape(-1, 2) for y in Y_hat_scaled], axis=0)
)


#计算残差作为LSTM的输入
# Delta_X_L,Y_hat_scaled这好像是一个二院数组捏妈妈的我咋知道他是个啥

epsilon_L_list = []
epsilon_R_list = []
for i in range(len(Y_hat)):
    epsilon_L = Delta_X_L[i+1] - Y_hat[i, 0]
    epsilon_R = Delta_X_R[i+1] - Y_hat[i, 1]
    epsilon_L_list.append(epsilon_L)
    epsilon_R_list.append(epsilon_R)
raw_data = pd.DataFrame({
    'high': epsilon_R_list,  # 上界数据
    'low': epsilon_L_list    # 下界数据
})

# 生成特征
raw_data['mid'] = (raw_data['high'] + raw_data['low']) / 2
raw_data['range'] = raw_data['high'] - raw_data['low']

# 按时间顺序划分原始数据集
train_size = int(len(raw_data) * 0.8)
train_data = raw_data.iloc[:train_size]
test_data = raw_data.iloc[train_size:]

# 标准化处理（仅在训练集上拟合）
scaler_mid = MinMaxScaler()
scaler_range = MinMaxScaler()

# 训练集标准化
train_mid_scaled = scaler_mid.fit_transform(train_data[['mid']])
train_range_scaled = scaler_range.fit_transform(train_data[['range']])

# 测试集转换
test_mid_scaled = scaler_mid.transform(test_data[['mid']])
test_range_scaled = scaler_range.transform(test_data[['range']])

# 合并标准化后的数据
full_mid = np.concatenate([train_mid_scaled, test_mid_scaled])
full_range = np.concatenate([train_range_scaled, test_range_scaled])


# 创建监督数据集
def create_dataset(mid, interval, look_back=10):
    X, y = [], []
    for i in range(len(mid) - look_back):
        X.append(np.column_stack((mid[i:i + look_back], interval[i:i + look_back])))
        y.append([mid[i + look_back], interval[i + look_back]])
    return np.array(X), np.array(y)


look_back = 10
X_full, y_full = create_dataset(full_mid, full_range, look_back)

# 正确划分训练测试集（考虑look_back偏移）
X_train = X_full
y_train = y_full
# X_train = X_full[:train_size - look_back]
# y_train = y_full[:train_size - look_back]
X_test = X_full[train_size - look_back:]
y_test = y_full[train_size - look_back:]

# 构建LSTM模型
model = Sequential([
    LSTM(64, input_shape=(look_back, 2), return_sequences=False),
    Dense(32, activation='relu'),
    Dense(2)
])
model.compile(optimizer='adam', loss='mse')
model.summary()

# 训练模型
history = model.fit(
    X_train, y_train,
    epochs=50,
    batch_size=32,
    validation_data=(X_test, y_test),
    verbose=1
)

# 全量预测
y_pred_scaled = model.predict(X_full)

# 反标准化
def inverse_scale(data, scaler):
    return scaler.inverse_transform(data.reshape(-1, 1)).flatten()


# 处理预测结果
pred_mid = inverse_scale(y_pred_scaled[:, 0], scaler_mid)
pred_range = inverse_scale(y_pred_scaled[:, 1], scaler_range)
pred_low = pred_mid - pred_range / 2
pred_high = pred_mid + pred_range / 2
Delta_X_L_hat_final_list = []
Delta_X_R_hat_final_list = []
for i in range(len(pred_low)):
    Delta_X_L_hat_final = pred_low[i] + Y_hat[i, 0]
    Delta_X_R_hat_final = pred_high[i] - Y_hat[i, 1]
    Delta_X_L_hat_final_list.append(Delta_X_L_hat_final)
    Delta_X_R_hat_final_list.append(Delta_X_R_hat_final)
XL_hat_final_list = []
XR_hat_final_list = []

for i in range(len(X_L)):
    if i <11:
        XL_hat_final = X_L[i]
        XR_hat_final = X_R[i]
    else:
        XL_hat_final = Delta_X_L_hat_final_list[i-12] + X_L[i - 1]
        XR_hat_final = Delta_X_R_hat_final_list[i-12] + X_R[i - 1]
    XL_hat_final_list.append(XL_hat_final)
    XR_hat_final_list.append(XR_hat_final)

def calculate_interval_coverage(X_L, X_R,X_hat_L,X_hat_R):
    Cover_rate_list = []
    for i in range(len(X_L)):
        if (X_L[i] < X_hat_R[i])and (X_hat_R[i] < X_R[i]):
            Cover_rate = (X_hat_R[i] - X_L[i]) / (X_R[i] - X_L[i])
        if (X_hat_L[i] < X_R[i]) and (X_R[i] < X_hat_R[i]):
            Cover_rate = (X_R[i] - X_hat_L[i]) / (X_R[i] - X_L[i])
        else:
            Cover_rate = 0
        Cover_rate_list.append(Cover_rate)
    mean_Cover_rate= np.mean(Cover_rate_list)
    return mean_Cover_rate
mean_Cover_rate = calculate_interval_coverage(X_L, X_R,XL_hat_final_list,XR_hat_final_list)
print(mean_Cover_rate)
# --------------------- 可视化结果 ---------------------
plt.figure(figsize=(18, 8))

# 创建时间索引（假设每日数据）
dates = pd.date_range(start='2023-01-01', periods=len(X_L), freq='D')  # 根据实际修改起始日期

# 绘制原始边界（带半透明效果）
plt.plot(dates, X_L, color='#3498db', alpha=0.8, linewidth=1.2, label='真实左边界')
plt.plot(dates, X_R, color='#e74c3c', alpha=0.8, linewidth=1.2, label='真实右边界')

# 绘制预测边界（带虚线样式）
plt.plot(dates, XL_hat_final_list, color='#2c3e50', linestyle='--', linewidth=1.5, label='预测左边界')
plt.plot(dates, XR_hat_final_list, color='#c0392b', linestyle='--', linewidth=1.5, label='预测右边界')

# 标记训练/测试分界
split_date = dates[int(len(dates)*0.8)]
plt.axvline(split_date, color='#7f8c8d', linestyle=':', linewidth=2, label='训练/测试分界')

# 添加填充区域（最后20%作为测试集）
test_start = int(len(dates)*0.8)
plt.fill_between(dates[test_start:],
                 XL_hat_final_list[test_start:],
                 XR_hat_final_list[test_start:],
                 color='#C0392B',
                 alpha=0.1,
                 label='测试集预测区间')
# 增强可视化配置
plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
plt.gca().xaxis.set_major_locator(mdates.MonthLocator(interval=1))
plt.xticks(rotation=45, ha='right')
plt.title('黄金期货边界预测对比\n模型覆盖度：{:.2%}'.format(mean_Cover_rate), fontsize=14, pad=20)
plt.ylabel('价格', fontsize=12)
plt.grid(alpha=0.3, linestyle='--')
plt.legend(loc='upper left', frameon=True, facecolor='white')

# 添加数据游标提示（可选）
def format_coord(x, y):
    try:
        idx = np.where(dates == pd.Timestamp(x))[0][0]
        return f'日期：{dates[idx]:%Y-%m-%d}  左边界：{X_L[idx]:.2f}  右边界：{X_R[idx]:.2f}'
    except:
        return f'X: {x:%Y-%m-%d}  Y: {y:.2f}'
plt.gca().format_coord = format_coord

plt.tight_layout()
plt.show()

# 控制台输出关键指标
print(f'''
[模型性能报告]
训练周期数：{len(history.history["loss"])}
最终训练损失：{history.history["loss"][-1]:.4e}
最终验证损失：{history.history["val_loss"][-1]:.4e}
区间覆盖指标：{mean_Cover_rate:.2%}
''')